import os
import csv
from openpyxl import Workbook, load_workbook

# --------------------------------------------------
# PATHS
# --------------------------------------------------
OUTPUT_DIR = "output"
EXCEL_FILE = os.path.join(OUTPUT_DIR, "profiles.xlsx")
CSV_FILE = os.path.join(OUTPUT_DIR, "profiles.csv")

HEADERS = [
    "username",
    "profile_url",
    "followers",
    "following",
    "posts",
    "category"
]

os.makedirs(OUTPUT_DIR, exist_ok=True)


# --------------------------------------------------
# UTILS
# --------------------------------------------------
def normalize_username(u):
    return (u or "").strip().lower()


# --------------------------------------------------
# MAIN SAVE FUNCTION
# --------------------------------------------------
def save_profile_excel(data):
    username = normalize_username(data.get("username"))
    if not username:
        return False

    # ==================================================
    # ----------- EXCEL HANDLING -----------------------
    # ==================================================
    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = "Profiles"
        ws.append(HEADERS)
        wb.save(EXCEL_FILE)

    wb = load_workbook(EXCEL_FILE)
    ws = wb.active

    existing = set()
    for row in ws.iter_rows(min_row=2, max_col=1):
        if row[0].value:
            existing.add(normalize_username(row[0].value))

    if username in existing:
        return False  # ✅ real duplicate

    ws.append([
        data.get("username"),
        data.get("profile_url"),
        data.get("followers"),
        data.get("following"),
        data.get("posts"),
        data.get("category"),
    ])

    wb.save(EXCEL_FILE)

    # ==================================================
    # ----------- CSV HANDLING -------------------------
    # ==================================================
    file_exists = os.path.exists(CSV_FILE)

    with open(CSV_FILE, "a", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)

        # write header only once
        if not file_exists:
            writer.writerow(HEADERS)

        writer.writerow([
            data.get("username"),
            data.get("profile_url"),
            data.get("followers"),
            data.get("following"),
            data.get("posts"),
            data.get("category"),
        ])

    return True
